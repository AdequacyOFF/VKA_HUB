"""Moderator router"""

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from sqlalchemy.orm import selectinload
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from io import BytesIO

from app.presentation.api.dependencies import get_db, require_moderator
from app.presentation.api.dtos.moderator import (
    AssignModeratorRequest,
    RemoveModeratorRequest,
    ModeratorResponse,
    UpdateUserLoginRequest,          # ✅ Используем созданные DTO
    ResetUserPasswordRequest,        # ✅
    UpdateControlQuestionRequest,    # ✅
    BanUserRequest,                  # ✅
    UnbanUserRequest,                # ✅
    UserSecurityInfoResponse,
    RespondToPlatformComplaintRequest
)
from app.infrastructure.repositories.user_repository_impl import UserRepositoryImpl
from app.infrastructure.repositories.moderator_repository_impl import ModeratorRepositoryImpl
from app.infrastructure.security.password import hash_password
from app.domain.models.user import User
from app.domain.models.moderator import Moderator
from app.domain.models.team import Team
from app.domain.models.competition import Competition
from app.domain.models.moderator_report import ModeratorReport
from app.domain.models.log import Log, ActivityType

router = APIRouter(prefix="/api/moderator", tags=["Moderator"])

# ✅ 1. Обновление логина пользователя
@router.patch("/users/{user_id}/login")
async def moderator_update_user_login(
    user_id: int,
    request: UpdateUserLoginRequest,  # ✅ Используем DTO
    current_user = Depends(require_moderator),
    db: AsyncSession = Depends(get_db)
):
    """Update user login (moderator only)"""
    user_repo = UserRepositoryImpl(db)
    
    # Получаем пользователя
    user = await user_repo.get_by_id(user_id)
    if not user:
        raise HTTPException(status_code=404, detail="Пользователь не найден")
    
    # Проверяем, не занят ли логин
    existing_user = await user_repo.get_by_login(request.login)
    if existing_user and existing_user.id != user_id:
        raise HTTPException(
            status_code=400, 
            detail="Логин уже занят другим пользователем"
        )
    
    # Обновляем логин
    try:
        updated_user = await user_repo.update(user_id, {"login": request.login})
        await db.commit()
        
        return {
            "message": "Логин пользователя успешно обновлен",
            "new_login": request.login,
            "user_id": user_id
        }
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка при обновлении логина: {str(e)}")

# ✅ 2. Сброс пароля пользователя
@router.post("/users/{user_id}/reset-password")
async def moderator_reset_user_password(
    user_id: int,
    request: ResetUserPasswordRequest,  # ✅ Используем DTO
    current_user = Depends(require_moderator),
    db: AsyncSession = Depends(get_db)
):
    """Reset user password (moderator only)"""
    user_repo = UserRepositoryImpl(db)
    
    # Получаем пользователя
    user = await user_repo.get_by_id(user_id)
    if not user:
        raise HTTPException(status_code=404, detail="Пользователь не найден")
    
    # Хешируем новый пароль
    hashed_password = hash_password(request.new_password)
    
    # Обновляем пароль
    try:
        updated_user = await user_repo.update(user_id, {"password_hash": hashed_password})
        await db.commit()
        
        return {
            "message": "Пароль пользователя успешно сброшен",
            "user_id": user_id
        }
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка при сбросе пароля: {str(e)}")

# ✅ 3. Обновление контрольного вопроса
@router.put("/users/{user_id}/control-question")
async def moderator_update_control_question(
    user_id: int,
    request: UpdateControlQuestionRequest,  # ✅ Используем DTO
    current_user = Depends(require_moderator),
    db: AsyncSession = Depends(get_db)
):
    """Update user control question and answer (moderator only)"""
    user_repo = UserRepositoryImpl(db)
    
    # Получаем пользователя
    user = await user_repo.get_by_id(user_id)
    if not user:
        raise HTTPException(status_code=404, detail="Пользователь не найден")
    
    # Хешируем ответ
    hashed_answer = hash_password(request.control_answer.lower().strip())
    
    # Обновляем контрольный вопрос и ответ
    try:
        updated_user = await user_repo.update(
            user_id, 
            {
                "control_question": request.control_question,
                "control_answer_hash": hashed_answer
            }
        )
        await db.commit()
        
        return {
            "message": "Контрольный вопрос пользователя успешно обновлен",
            "user_id": user_id
        }
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка при обновлении контрольного вопроса: {str(e)}")

# ✅ 4. Блокировка пользователя (с поддержкой причины)
@router.post("/users/{user_id}/ban")
async def moderator_ban_user(
    user_id: int,
    request: BanUserRequest = None,  # ✅ DTO с опциональной причиной
    current_user = Depends(require_moderator),
    db: AsyncSession = Depends(get_db)
):
    """Ban user (moderator only)"""
    user_repo = UserRepositoryImpl(db)
    
    # Получаем пользователя
    user = await user_repo.get_by_id(user_id)
    if not user:
        raise HTTPException(status_code=404, detail="Пользователь не найден")
    
    # Проверяем, не пытаемся ли заблокировать себя
    if user.id == current_user.id:
        raise HTTPException(status_code=400, detail="Нельзя заблокировать самого себя")
    
    # Проверяем, не пытаемся ли заблокировать системного пользователя
    if hasattr(user, 'login') and user.login == "GeDeKo":
        raise HTTPException(status_code=403, detail="Нельзя заблокировать системного пользователя")
    
    # Проверяем, не пытаемся ли заблокировать другого модератора
    is_user_moderator = await db.execute(
        select(Moderator).where(Moderator.user_id == user_id)
    )
    if is_user_moderator.scalar():
        raise HTTPException(
            status_code=400, 
            detail="Нельзя заблокировать модератора. Сначала снимите права модератора."
        )
    
    # Блокируем пользователя
    try:
        updated_user = await user_repo.update(user_id, {"is_banned": True})
        await db.commit()
        
        response = {
            "message": "Пользователь успешно заблокирован",
            "user_id": user_id,
        }
        
        # Log the ban action
        reason = request.reason if request and request.reason else "Причина не указана"
        log_entry = Log(
            user_id=current_user.id,
            action_type=ActivityType.USER_BAN,
            description=f"Заблокирован пользователь {user.login} (ID: {user_id})",
            action_metadata={
                "target_user_id": user_id,
                "target_login": user.login,
                "reason": reason
            }
        )
        db.add(log_entry)

        # Добавляем причину в ответ если есть
        if request and request.reason:
            response["reason"] = request.reason

        return response
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка при блокировке пользователя: {str(e)}")

# ✅ 5. Разблокировка пользователя (с поддержкой причины)
@router.post("/users/{user_id}/unban")
async def moderator_unban_user(
    user_id: int,
    request: UnbanUserRequest = None,  # ✅ DTO с опциональной причиной
    current_user = Depends(require_moderator),
    db: AsyncSession = Depends(get_db)
):
    """Unban user (moderator only)"""
    user_repo = UserRepositoryImpl(db)
    
    # Получаем пользователя
    user = await user_repo.get_by_id(user_id)
    if not user:
        raise HTTPException(status_code=404, detail="Пользователь не найден")
    
    # Разблокируем пользователя
    try:
        updated_user = await user_repo.update(user_id, {"is_banned": False})
        await db.commit()
        
        response = {
            "message": "Пользователь успешно разблокирован",
            "user_id": user_id,
        }

        # Log the unban action
        reason = request.reason if request and request.reason else "Причина не указана"
        log_entry = Log(
            user_id=current_user.id,
            action_type=ActivityType.USER_UNBAN,
            description=f"Разблокирован пользователь {user.login} (ID: {user_id})",
            action_metadata={
                "target_user_id": user_id,
                "target_login": user.login,
                "reason": reason
            }
        )
        db.add(log_entry)

        # Добавляем причину в ответ если есть
        if request and request.reason:
            response["reason"] = request.reason

        return response
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка при разблокировке пользователя: {str(e)}")

# ✅ 6. Получение полной информации о пользователе для модератора
@router.get("/users/{user_id}")
async def moderator_get_user(
    user_id: int,
    current_user = Depends(require_moderator),
    db: AsyncSession = Depends(get_db)
):
    """Get detailed user information for moderator"""
    user_repo = UserRepositoryImpl(db)
    
    # Получаем пользователя
    user = await user_repo.get_by_id(user_id)
    if not user:
        raise HTTPException(status_code=404, detail="Пользователь не найден")
    
    # Проверяем, является ли пользователь модератором
    is_moderator_result = await db.execute(
        select(Moderator).where(Moderator.user_id == user_id)
    )
    is_moderator = is_moderator_result.scalar() is not None
    
    # Формируем ответ
    response = {
        "id": user.id,
        "login": user.login,
        "first_name": user.first_name,
        "last_name": user.last_name,
        "middle_name": user.middle_name,
        "study_group": user.study_group,
        "position": user.position,
        "rank": user.rank,
        "avatar_url": user.avatar_url,
        "control_question": user.control_question,
        "is_banned": getattr(user, 'is_banned', False),
        "is_moderator": is_moderator,
        "created_at": user.created_at,
        "updated_at": user.updated_at,
    }
    
    return response

# ✅ 7. Поиск пользователей с фильтрами
@router.get("/users")
async def moderator_search_users(
    search: str = None,
    study_group: str = None,
    is_banned: bool = None,
    skip: int = 0,
    limit: int = 100,
    current_user = Depends(require_moderator),
    db: AsyncSession = Depends(get_db)
):
    """Search users with filters (moderator only)"""
    user_repo = UserRepositoryImpl(db)
    
    # Собираем фильтры
    filters = {}
    if search:
        filters["search"] = search
    if study_group:
        filters["study_group"] = study_group
    if is_banned is not None:
        # Предполагаем, что модель User имеет поле is_banned
        filters["is_banned"] = is_banned
    
    # Получаем пользователей
    users = await user_repo.list_users(skip, limit, filters)
    total = await user_repo.count_users(filters)
    
    # Формируем ответ
    return {
        "items": [
            {
                "id": user.id,
                "login": user.login,
                "first_name": user.first_name,
                "last_name": user.last_name,
                "study_group": user.study_group,
                "is_banned": getattr(user, 'is_banned', False),
                "created_at": user.created_at
            }
            for user in users
        ],
        "total": total,
        "skip": skip,
        "limit": limit
    }

@router.get("/users/{user_id}/security")
async def get_user_security_info(
    user_id: int,
    current_user = Depends(require_moderator),
    db: AsyncSession = Depends(get_db)
):
    """Get user's sensitive security information (moderators only)"""
    user_repo = UserRepositoryImpl(db)
    user = await user_repo.get_by_id(user_id)
    
    if not user:
        raise HTTPException(status_code=404, detail="Пользователь не найден")
    
    # Получаем дополнительную информацию о пользователе
    # Проверяем, является ли пользователь модератором
    is_moderator_result = await db.execute(
        select(Moderator).where(Moderator.user_id == user_id)
    )
    is_moderator = is_moderator_result.scalar() is not None
    
    # Определяем дату последнего изменения пароля
    # Если нет отдельного поля, используем updated_at
    password_last_changed = user.updated_at
    
    # Определяем дату настройки безопасности
    security_setup_date = user.updated_at if user.control_question else None
    
    return UserSecurityInfoResponse(
        id=user.id,
        login=user.login,
        control_question=user.control_question,
        has_control_answer=bool(user.control_answer_hash),
        password_last_changed=password_last_changed,
        security_setup_date=security_setup_date,
        # Дополнительная информация для контекста
        is_moderator=is_moderator,
        is_banned=getattr(user, 'is_banned', False)
    )


# ========================================
# NEW ENDPOINTS - List, Stats, Report Generation
# ========================================

@router.get("/list")
async def list_moderators(
    current_user = Depends(require_moderator),
    db: AsyncSession = Depends(get_db)
):
    """List all moderators with user details"""
    # Get all moderators with their user information
    result = await db.execute(
        select(Moderator, User)
        .join(User, Moderator.user_id == User.id)
        .order_by(Moderator.assigned_at.desc())
    )

    moderators = []
    for mod, user in result.all():
        moderators.append({
            "id": mod.id,
            "user_id": user.id,
            "login": user.login,
            "first_name": user.first_name,
            "last_name": user.last_name,
            "middle_name": user.middle_name,
            "avatar_url": user.avatar_url,
            "assigned_at": mod.assigned_at,
            "assigned_by": mod.assigned_by
        })

    return {
        "items": moderators,
        "total": len(moderators)
    }


@router.get("/stats")
async def get_platform_stats(
    current_user = Depends(require_moderator),
    db: AsyncSession = Depends(get_db)
):
    """Get platform statistics for moderator dashboard"""
    # Calculate the date range for "this month"
    now = datetime.utcnow()
    month_ago = now - relativedelta(months=1)
    two_months_ago = now - relativedelta(months=2)

    # Total counts
    total_users = await db.scalar(select(func.count(User.id)))
    total_teams = await db.scalar(select(func.count(Team.id)))
    total_competitions = await db.scalar(select(func.count(Competition.id)))

    # Active competitions (not ended)
    active_competitions = await db.scalar(
        select(func.count(Competition.id))
        .where(Competition.end_date >= now.date())
    )

    # New users this month
    new_users_this_month = await db.scalar(
        select(func.count(User.id))
        .where(User.created_at >= month_ago)
    )

    # New users previous month (for growth calculation)
    new_users_prev_month = await db.scalar(
        select(func.count(User.id))
        .where(User.created_at >= two_months_ago)
        .where(User.created_at < month_ago)
    )

    # New teams this month
    new_teams_this_month = await db.scalar(
        select(func.count(Team.id))
        .where(Team.created_at >= month_ago)
    )

    # New teams previous month
    new_teams_prev_month = await db.scalar(
        select(func.count(Team.id))
        .where(Team.created_at >= two_months_ago)
        .where(Team.created_at < month_ago)
    )

    # Calculate growth percentages
    def calc_growth(current, previous):
        if previous == 0:
            return 100.0 if current > 0 else 0.0
        return round(((current - previous) / previous) * 100, 1)

    user_growth = calc_growth(new_users_this_month, new_users_prev_month)
    team_growth = calc_growth(new_teams_this_month, new_teams_prev_month)

    # Reports stats - get from user_complaints table
    from app.domain.models.user_complaint import UserComplaint, ComplaintStatus
    from app.domain.models.platform_complaint import PlatformComplaint, ComplaintStatus as PlatformComplaintStatus

    pending_reports = await db.scalar(
        select(func.count(UserComplaint.id))
        .where(UserComplaint.status == ComplaintStatus.PENDING)
    )

    resolved_reports = await db.scalar(
        select(func.count(UserComplaint.id))
        .where(UserComplaint.status == ComplaintStatus.RESOLVED)
    )

    rejected_reports = await db.scalar(
        select(func.count(UserComplaint.id))
        .where(UserComplaint.status == ComplaintStatus.REJECTED)
    )

    # Platform complaints (feedback) stats
    pending_feedback = await db.scalar(
        select(func.count(PlatformComplaint.id))
        .where(PlatformComplaint.status == PlatformComplaintStatus.PENDING)
    )

    resolved_feedback = await db.scalar(
        select(func.count(PlatformComplaint.id))
        .where(PlatformComplaint.status == PlatformComplaintStatus.RESOLVED)
    )

    rejected_feedback = await db.scalar(
        select(func.count(PlatformComplaint.id))
        .where(PlatformComplaint.status == PlatformComplaintStatus.REJECTED)
    )

    return {
        "total_users": total_users or 0,
        "total_teams": total_teams or 0,
        "total_competitions": total_competitions or 0,
        "active_competitions": active_competitions or 0,
        "pending_reports": pending_reports or 0,
        "resolved_reports": resolved_reports or 0,
        "rejected_reports": rejected_reports or 0,
        "pending_feedback": pending_feedback or 0,
        "resolved_feedback": resolved_feedback or 0,
        "rejected_feedback": rejected_feedback or 0,
        "new_users_this_month": new_users_this_month or 0,
        "new_teams_this_month": new_teams_this_month or 0,
        "user_growth": user_growth,
        "team_growth": team_growth
    }


@router.post("/reports/generate")
async def generate_competition_report(
    competition_id: int,
    current_user = Depends(require_moderator),
    db: AsyncSession = Depends(get_db)
):
    """Generate report for a competition"""
    from app.infrastructure.storage.report_generator import generate_competition_report_docx
    from app.infrastructure.repositories.competition_repository_impl import CompetitionRepositoryImpl

    # Get competition
    comp_repo = CompetitionRepositoryImpl(db)
    competition = await comp_repo.get_by_id(competition_id)

    if not competition:
        raise HTTPException(status_code=404, detail="Соревнование не найдено")

    try:
        # Generate the report file
        file_url = await generate_competition_report_docx(competition, db)

        # Save report record
        report = ModeratorReport(
            competition_id=competition_id,
            generated_by=current_user.id,
            file_url=file_url
        )
        db.add(report)
        await db.commit()
        await db.refresh(report)

        return {
            "message": "Отчет успешно сгенерирован",
            "report_id": report.id,
            "file_url": file_url,
            "generated_at": report.generated_at
        }
    except Exception as e:
        await db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Ошибка при генерации отчета: {str(e)}"
        )


@router.post("/assign")
async def assign_moderator(
    request: AssignModeratorRequest,
    current_user = Depends(require_moderator),
    db: AsyncSession = Depends(get_db)
):
    """Assign moderator role to a user"""
    mod_repo = ModeratorRepositoryImpl(db)
    user_repo = UserRepositoryImpl(db)

    # Check if user exists
    user = await user_repo.get_by_id(request.user_id)
    if not user:
        raise HTTPException(status_code=404, detail="Пользователь не найден")

    # Check if already a moderator
    if await mod_repo.is_moderator(request.user_id):
        raise HTTPException(status_code=400, detail="Пользователь уже является модератором")

    try:
        await mod_repo.assign(request.user_id, current_user.id)
        await db.commit()

        return {
            "message": "Права модератора успешно назначены",
            "user_id": request.user_id
        }
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка: {str(e)}")


@router.post("/remove")
async def remove_moderator(
    request: RemoveModeratorRequest,
    current_user = Depends(require_moderator),
    db: AsyncSession = Depends(get_db)
):
    """Remove moderator role from a user"""
    mod_repo = ModeratorRepositoryImpl(db)

    # Check if user is a moderator
    if not await mod_repo.is_moderator(request.user_id):
        raise HTTPException(status_code=400, detail="Пользователь не является модератором")

    # Prevent removing yourself
    if request.user_id == current_user.id:
        raise HTTPException(status_code=400, detail="Нельзя снять права модератора с самого себя")

    try:
        await mod_repo.remove(request.user_id)
        await db.commit()

        return {
            "message": "Права модератора успешно сняты",
            "user_id": request.user_id
        }
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка: {str(e)}")


# ========================================
# USER COMPLAINTS ENDPOINTS
# ========================================

@router.get("/reports")
async def list_user_complaints(
    skip: int = 0,
    limit: int = 100,
    status: str = None,
    current_user = Depends(require_moderator),
    db: AsyncSession = Depends(get_db)
):
    """List all user complaints for moderation"""
    from app.infrastructure.repositories.user_complaint_repository_impl import UserComplaintRepositoryImpl
    from app.presentation.api.dtos.moderator import UserComplaintResponse

    complaint_repo = UserComplaintRepositoryImpl(db)

    complaints = await complaint_repo.list_all(skip=skip, limit=limit, status=status)
    total = await complaint_repo.count(status=status)

    # Build response with reporter/target logins
    items = []
    for complaint in complaints:
        # Get reporter and target logins
        reporter = await db.execute(select(User).where(User.id == complaint.reporter_id))
        reporter_user = reporter.scalar_one_or_none()
        target = await db.execute(select(User).where(User.id == complaint.target_id))
        target_user = target.scalar_one_or_none()

        items.append(UserComplaintResponse(
            id=complaint.id,
            reporter=reporter_user.login if reporter_user else f"User #{complaint.reporter_id}",
            target=target_user.login if target_user else f"User #{complaint.target_id}",
            reason=complaint.reason,
            description=complaint.description,
            status=complaint.status.value,
            created_at=complaint.created_at,
            updated_at=complaint.updated_at
        ))

    return {
        "items": items,
        "total": total
    }


@router.post("/reports/{complaint_id}/approve")
async def approve_complaint(
    complaint_id: int,
    current_user = Depends(require_moderator),
    db: AsyncSession = Depends(get_db)
):
    """Approve a user complaint"""
    from app.infrastructure.repositories.user_complaint_repository_impl import UserComplaintRepositoryImpl
    from app.domain.models.user_complaint import ComplaintStatus

    complaint_repo = UserComplaintRepositoryImpl(db)

    complaint = await complaint_repo.get_by_id(complaint_id)
    if not complaint:
        raise HTTPException(status_code=404, detail="Жалоба не найдена")

    if complaint.status != ComplaintStatus.PENDING:
        raise HTTPException(status_code=400, detail="Жалоба уже рассмотрена")

    try:
        await complaint_repo.update_status(complaint_id, ComplaintStatus.RESOLVED, current_user.id)
        await db.commit()

        return {
            "message": "Жалоба одобрена",
            "complaint_id": complaint_id
        }
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка: {str(e)}")


@router.post("/reports/{complaint_id}/reject")
async def reject_complaint(
    complaint_id: int,
    current_user = Depends(require_moderator),
    db: AsyncSession = Depends(get_db)
):
    """Reject a user complaint"""
    from app.infrastructure.repositories.user_complaint_repository_impl import UserComplaintRepositoryImpl
    from app.domain.models.user_complaint import ComplaintStatus

    complaint_repo = UserComplaintRepositoryImpl(db)

    complaint = await complaint_repo.get_by_id(complaint_id)
    if not complaint:
        raise HTTPException(status_code=404, detail="Жалоба не найдена")

    if complaint.status != ComplaintStatus.PENDING:
        raise HTTPException(status_code=400, detail="Жалоба уже рассмотрена")

    try:
        await complaint_repo.update_status(complaint_id, ComplaintStatus.REJECTED, current_user.id)
        await db.commit()

        return {
            "message": "Жалоба отклонена",
            "complaint_id": complaint_id
        }
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка: {str(e)}")


# ========================================
# PLATFORM COMPLAINTS (Site Feedback) ENDPOINTS
# ========================================

@router.get("/platform-complaints")
async def list_platform_complaints(
    skip: int = 0,
    limit: int = 100,
    status: str = None,
    category: str = None,
    current_user = Depends(require_moderator),
    db: AsyncSession = Depends(get_db)
):
    """List all platform complaints for moderation"""
    from app.infrastructure.repositories.platform_complaint_repository_impl import PlatformComplaintRepositoryImpl
    from app.presentation.api.dtos.moderator import PlatformComplaintResponse

    complaint_repo = PlatformComplaintRepositoryImpl(db)

    complaints = await complaint_repo.list_all(skip=skip, limit=limit, status=status, category=category)
    total = await complaint_repo.count(status=status, category=category)

    # Build response with user logins
    items = []
    for complaint in complaints:
        # Get user login
        user = await db.execute(select(User).where(User.id == complaint.user_id))
        user_obj = user.scalar_one_or_none()

        items.append(PlatformComplaintResponse(
            id=complaint.id,
            user_id=complaint.user_id,
            user=user_obj.login if user_obj else f"User #{complaint.user_id}",
            category=complaint.category.value,
            priority=complaint.priority.value,
            title=complaint.title,
            description=complaint.description,
            status=complaint.status.value,
            moderator_response=complaint.moderator_response,
            response_read=complaint.response_read,
            resolved_by=complaint.resolved_by,
            created_at=complaint.created_at,
            updated_at=complaint.updated_at
        ))

    return {
        "items": items,
        "total": total
    }


@router.post("/platform-complaints/{complaint_id}/respond")
async def respond_to_platform_complaint(
    complaint_id: int,
    request: RespondToPlatformComplaintRequest,
    current_user = Depends(require_moderator),
    db: AsyncSession = Depends(get_db)
):
    """Respond to a platform complaint with moderator feedback"""
    from app.infrastructure.repositories.platform_complaint_repository_impl import PlatformComplaintRepositoryImpl
    from app.domain.models.platform_complaint import ComplaintStatus

    complaint_repo = PlatformComplaintRepositoryImpl(db)

    complaint = await complaint_repo.get_by_id(complaint_id)
    if not complaint:
        raise HTTPException(status_code=404, detail="Обращение не найдено")

    if complaint.status != ComplaintStatus.PENDING:
        raise HTTPException(status_code=400, detail="Обращение уже рассмотрено")

    # Validate status
    if request.status not in ['resolved', 'rejected']:
        raise HTTPException(status_code=400, detail="Недопустимый статус. Используйте 'resolved' или 'rejected'")

    try:
        status_enum = ComplaintStatus.RESOLVED if request.status == 'resolved' else ComplaintStatus.REJECTED
        await complaint_repo.respond_to_complaint(
            complaint_id,
            request.response,
            status_enum,
            current_user.id
        )
        await db.commit()

        return {
            "message": "Ответ отправлен",
            "complaint_id": complaint_id,
            "status": request.status
        }
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка: {str(e)}")


# ========================================
# ANALYTICS ENDPOINT
# ========================================

@router.get("/analytics")
async def get_platform_analytics(
    current_user = Depends(require_moderator),
    db: AsyncSession = Depends(get_db)
):
    """Get platform analytics data for charts"""
    from calendar import month_abbr
    from app.presentation.api.dtos.moderator import AnalyticsResponse

    # Russian month names
    month_names = {
        1: "Янв", 2: "Фев", 3: "Мар", 4: "Апр", 5: "Май", 6: "Июн",
        7: "Июл", 8: "Авг", 9: "Сен", 10: "Окт", 11: "Ноя", 12: "Дек"
    }

    now = datetime.utcnow()

    # Get user growth data for the last 6 months
    user_growth = []
    for i in range(5, -1, -1):
        target_date = now - relativedelta(months=i)
        month_start = target_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        month_end = (month_start + relativedelta(months=1))

        count = await db.scalar(
            select(func.count(User.id))
            .where(User.created_at >= month_start)
            .where(User.created_at < month_end)
        )

        user_growth.append({
            "month": month_names[target_date.month],
            "users": count or 0
        })

    # Get team growth data for the last 6 months
    team_stats = []
    for i in range(5, -1, -1):
        target_date = now - relativedelta(months=i)
        month_start = target_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        month_end = (month_start + relativedelta(months=1))

        count = await db.scalar(
            select(func.count(Team.id))
            .where(Team.created_at >= month_start)
            .where(Team.created_at < month_end)
        )

        team_stats.append({
            "month": month_names[target_date.month],
            "teams": count or 0
        })

    # Get competition types distribution
    competition_types = []
    type_result = await db.execute(
        select(Competition.type, func.count(Competition.id))
        .group_by(Competition.type)
    )

    type_names = {
        "hackathon": "Хакатоны",
        "olympiad": "Олимпиады",
        "championship": "Чемпионаты",
        "ctf": "CTF"
    }

    for comp_type, count in type_result.all():
        type_value = comp_type.value if hasattr(comp_type, 'value') else str(comp_type)
        competition_types.append({
            "name": type_names.get(type_value, type_value),
            "value": count
        })

    return {
        "userGrowth": user_growth,
        "teamStats": team_stats,
        "competitionTypes": competition_types
    }


@router.get("/analytics/export")
async def export_prize_winning_report(
    period: str,
    current_user = Depends(require_moderator),
    db: AsyncSession = Depends(get_db)
):
    """Export prize-winning competition results to Excel

    Generates an Excel file with prize-winning teams (1st, 2nd, 3rd place).

    Columns:
    - No. (serial number)
    - Competition name
    - Competition organizer
    - Competition date
    - Place occupied
    - Team members data (serial number, rank, full name, study group)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from app.domain.models.competition_report import CompetitionReport, CompetitionResult
    from app.domain.models.competition_registration import CompetitionRegistration
    from app.domain.models.competition_team_member import CompetitionTeamMember

    # Calculate date range based on period
    now = datetime.utcnow()
    if period == "week":
        start_date = now - timedelta(days=7)
    elif period == "month":
        start_date = now - relativedelta(months=1)
    elif period == "quarter":
        start_date = now - relativedelta(months=3)
    elif period == "year":
        start_date = now - relativedelta(years=1)
    else:
        raise HTTPException(status_code=400, detail="Недопустимый период. Используйте: week, month, quarter, year")

    # Query prize-winning reports (1st, 2nd, 3rd place only)
    prize_results = [CompetitionResult.FIRST_PLACE, CompetitionResult.SECOND_PLACE, CompetitionResult.THIRD_PLACE]

    result = await db.execute(
        select(CompetitionReport)
        .join(CompetitionRegistration, CompetitionReport.registration_id == CompetitionRegistration.id)
        .join(Competition, CompetitionRegistration.competition_id == Competition.id)
        .options(
            selectinload(CompetitionReport.registration)
            .selectinload(CompetitionRegistration.competition),
            selectinload(CompetitionReport.registration)
            .selectinload(CompetitionRegistration.team_members)
            .selectinload(CompetitionTeamMember.user)
        )
        .where(CompetitionReport.result.in_(prize_results))
        .where(CompetitionReport.submitted_at >= start_date)
        .order_by(CompetitionReport.submitted_at.desc())
    )

    reports = result.scalars().all()

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Призовые места"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563B8", end_color="2563B8", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Create headers
    headers = [
        "№ п/п",
        "Наименование соревнования",
        "Организатор соревнования",
        "Дата соревнования",
        "Занятое место",
        "Данные участников команды"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Set column widths
    ws.column_dimensions['A'].width = 8   # No.
    ws.column_dimensions['B'].width = 35  # Competition name
    ws.column_dimensions['C'].width = 30  # Organizer
    ws.column_dimensions['D'].width = 18  # Date
    ws.column_dimensions['E'].width = 15  # Place
    ws.column_dimensions['F'].width = 60  # Team members

    # Map result enum to Russian text
    result_map = {
        CompetitionResult.FIRST_PLACE: "1 место",
        CompetitionResult.SECOND_PLACE: "2 место",
        CompetitionResult.THIRD_PLACE: "3 место"
    }

    # Fill data rows
    for row_num, report in enumerate(reports, 2):
        registration = report.registration
        competition = registration.competition

        # Format competition date
        date_str = competition.start_date.strftime("%d.%m.%Y") if competition.start_date else ""
        if competition.end_date and competition.end_date != competition.start_date:
            date_str += f" - {competition.end_date.strftime('%d.%m.%Y')}"

        # Format team members data
        members_data = []
        for idx, member in enumerate(registration.team_members, 1):
            user = member.user
            if user:
                # Full name: Last First Middle
                full_name = f"{user.last_name or ''} {user.first_name or ''} {user.middle_name or ''}".strip()
                rank = user.rank or "-"
                study_group = user.study_group or "-"
                members_data.append(f"{idx}. {rank}, {full_name}, гр. {study_group}")

        members_str = "\n".join(members_data) if members_data else "-"

        # Write row data
        ws.cell(row=row_num, column=1, value=row_num - 1).alignment = cell_alignment
        ws.cell(row=row_num, column=1).border = thin_border

        ws.cell(row=row_num, column=2, value=competition.name).alignment = cell_alignment
        ws.cell(row=row_num, column=2).border = thin_border

        ws.cell(row=row_num, column=3, value=competition.organizer or "-").alignment = cell_alignment
        ws.cell(row=row_num, column=3).border = thin_border

        ws.cell(row=row_num, column=4, value=date_str).alignment = cell_alignment
        ws.cell(row=row_num, column=4).border = thin_border

        ws.cell(row=row_num, column=5, value=result_map.get(report.result, str(report.result))).alignment = cell_alignment
        ws.cell(row=row_num, column=5).border = thin_border

        ws.cell(row=row_num, column=6, value=members_str).alignment = cell_alignment
        ws.cell(row=row_num, column=6).border = thin_border

        # Adjust row height based on number of members
        ws.row_dimensions[row_num].height = max(20, len(members_data) * 15)

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Generate filename with period info (ASCII-safe for headers)
    period_names = {
        "week": "week",
        "month": "month",
        "quarter": "quarter",
        "year": "year"
    }
    filename = f"prizovye_mesta_{period_names.get(period, period)}_{now.strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        }
    )